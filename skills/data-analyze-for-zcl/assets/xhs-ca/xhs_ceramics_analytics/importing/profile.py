from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

EXCEL_SUFFIXES = {".xls", ".xlsx"}


@dataclass(frozen=True)
class FileProfile:
    path: Path
    table_name: str
    columns: list[str]
    row_count: int
    sample_rows: list[dict[str, object]]
    sheet_name: str | None = None
    header_row: int | None = None
    confidence: float | None = None


def profile_csv(path: Path) -> FileProfile:
    frame = pd.read_csv(path)
    return profile_frame(
        path=path,
        table_name=path.stem,
        frame=frame,
    )


def profile_file(path: Path) -> FileProfile:
    if path.suffix.lower() in EXCEL_SUFFIXES:
        frame, sheet_name, header_row, confidence = load_excel_table(path)
        return profile_frame(
            path=path,
            table_name=sheet_name,
            frame=frame,
            sheet_name=sheet_name,
            header_row=header_row,
            confidence=confidence,
        )
    return profile_csv(path)


def load_table(path: Path, sheet: str | None = None) -> pd.DataFrame:
    if path.suffix.lower() in EXCEL_SUFFIXES:
        frame, _, _, _ = load_excel_table(path, sheet=sheet)
        return frame
    return pd.read_csv(path)


def load_excel_table(path: Path, sheet: str | None = None) -> tuple[pd.DataFrame, str, int, float]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = [sheet] if sheet else workbook.sheetnames
    candidates: list[tuple[float, float, int, str, int]] = []
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        header_row, raw_score, confidence, data_rows = _detect_excel_header(worksheet)
        candidates.append((raw_score, confidence, data_rows, sheet_name, header_row))
    if not candidates:
        raise ValueError(f"No readable sheets found in {path}")
    _, confidence, _, sheet_name, header_row = max(candidates, key=lambda item: (item[0], item[2]))
    frame = pd.read_excel(path, sheet_name=sheet_name, header=header_row, engine="openpyxl")
    frame = frame.dropna(axis=0, how="all").dropna(axis=1, how="all")
    frame.columns = _unique_columns(
        [_clean_column_name(column, index) for index, column in enumerate(frame.columns)]
    )
    return frame.reset_index(drop=True), sheet_name, header_row, confidence


def profile_frame(
    path: Path,
    table_name: str,
    frame: pd.DataFrame,
    sheet_name: str | None = None,
    header_row: int | None = None,
    confidence: float | None = None,
) -> FileProfile:
    sample = frame.head(5)
    sample_rows = sample.astype(object).where(pd.notna(sample), None).to_dict(orient="records")
    return FileProfile(
        path=path,
        table_name=table_name,
        columns=list(frame.columns),
        row_count=len(frame),
        sample_rows=sample_rows,
        sheet_name=sheet_name,
        header_row=header_row,
        confidence=confidence,
    )


def _detect_excel_header(worksheet) -> tuple[int, float, float, int]:
    rows = list(worksheet.iter_rows(values_only=True))
    non_empty_rows = [
        (index, row)
        for index, row in enumerate(rows)
        if any(value is not None and str(value).strip() for value in row)
    ]
    if not non_empty_rows:
        return 0, 0.0, 0.0, 0

    scan_rows = non_empty_rows[:20]
    scored_rows: list[tuple[float, int, int]] = []
    for index, row in scan_rows:
        cells = [value for value in row if value is not None and str(value).strip()]
        if not cells:
            continue
        text_cells = [value for value in cells if isinstance(value, str)]
        unique_text = len({str(value).strip() for value in text_cells})
        next_non_empty = _count_data_rows_after(rows, index)
        score = (len(cells) * 0.45) + (len(text_cells) * 0.35) + (unique_text * 0.15)
        if next_non_empty:
            score += 0.5
        scored_rows.append((score, index, next_non_empty))
    if not scored_rows:
        return 0, 0.0, 0.0, 0

    score, header_row, data_rows = max(scored_rows, key=lambda item: (item[0], item[2]))
    confidence = min(score / 5.0, 1.0)
    return header_row, score, confidence, data_rows


def _count_data_rows_after(rows: list[tuple[object, ...]], header_row: int) -> int:
    return sum(
        1
        for row in rows[header_row + 1 :]
        if any(value is not None and str(value).strip() for value in row)
    )


def _clean_column_name(value: object, index: int) -> str:
    if pd.isna(value):
        return f"column_{index + 1}"
    name = str(value).strip()
    return name or f"column_{index + 1}"


def _unique_columns(columns: list[str]) -> list[str]:
    used: dict[str, int] = {}
    unique: list[str] = []
    for column in columns:
        count = used.get(column, 0)
        if count:
            unique.append(f"{column}_{count + 1}")
        else:
            unique.append(column)
        used[column] = count + 1
    return unique
